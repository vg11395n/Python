# Task # 1 Write a program that searches for the businesses e.g. Best Barber Shop on Yelp

# install Pip package using command: pip install openpyxl

import openpyxl as xl


# from openpyxl import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    wb.save(filename)
